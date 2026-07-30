"""Extract selected province-year service-context indicators from the yearbook panel.

The source workbook is read-only. Output is a compact CSV for secondary,
ecological mechanism-consistency analyses. These indicators must not be used
as substitutes for individual service use or as post-treatment covariates in
the individual policy models.
"""

from __future__ import annotations

from pathlib import Path
import csv
import os
from openpyxl import load_workbook


source_value = os.environ.get("CHINA_HEALTH_YEARBOOK_PANEL")
output_value = os.environ.get("V7_OUTPUT_DIR")
if not source_value:
    raise RuntimeError("Set CHINA_HEALTH_YEARBOOK_PANEL to the panel workbook.")
if not output_value:
    raise RuntimeError("Set V7_OUTPUT_DIR to the analysis output directory.")
SOURCE = Path(source_value)
OUT_DIR = Path(output_value)

# Column numbers are verified against the workbook's "指标" sheet.
SELECTED = {
    32: "primary_care_institutions_count",
    211: "primary_care_staff_count",
    245: "township_staff_per_1000_agricultural_population",
    253: "village_clinic_staff_per_1000_agricultural_population",
    302: "primary_care_beds_count",
    384: "all_medical_visits_count",
    387: "health_examination_people_count",
    391: "visits_per_resident",
    399: "hospital_admissions_count",
    405: "annual_hospitalization_rate_percent",
    875: "total_population_10000",
    885: "population_age65plus_10000",
    888: "population_age65plus_percent",
}


def join_header(values: tuple[object, ...]) -> str:
    return " > ".join(str(v).strip() for v in values if v not in (None, ""))


def as_number(value: object) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def main() -> None:
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    panel = wb["卫生与区域经济"]
    indicator = wb["指标"]

    assert panel.max_row == 766, f"Unexpected panel rows: {panel.max_row}"
    assert panel.max_column == 913, f"Unexpected panel columns: {panel.max_column}"
    assert indicator.max_row == 924, f"Unexpected indicator rows: {indicator.max_row}"

    panel_rows = list(panel.iter_rows(min_row=1, max_row=766, values_only=True))
    indicator_rows = list(
        indicator.iter_rows(min_row=1, max_row=924, min_col=1, max_col=6, values_only=True)
    )
    assert panel_rows[6][0] == "安徽省" and panel_rows[6][1] == 2000
    assert panel_rows[-2][0] == "重庆市" and panel_rows[-2][1] == 2020

    audit_rows: list[dict[str, object]] = []
    for col, name in SELECTED.items():
        panel_header = join_header(tuple(panel_rows[r][col - 1] for r in range(6)))
        indicator_path = join_header(tuple(indicator_rows[col - 1]))
        assert panel_header == indicator_path, (
            f"Header mismatch at column {col}: {panel_header!r} != {indicator_path!r}"
        )
        audit_rows.append(
            {
                "source_column": col,
                "output_name": name,
                "indicator_path": indicator_path,
            }
        )

    output_rows: list[dict[str, object]] = []
    for row in panel_rows[6:]:
        area, year = row[0], row[1]
        if area in (None, "") or not isinstance(year, (int, float)):
            continue
        rec: dict[str, object] = {"area": area, "year": int(year)}
        for col, name in SELECTED.items():
            rec[name] = as_number(row[col - 1])

        pop = rec["total_population_10000"]
        if isinstance(pop, (int, float)) and pop > 0:
            rec["primary_care_institutions_per_10000"] = (
                rec["primary_care_institutions_count"] / pop
                if isinstance(rec["primary_care_institutions_count"], (int, float))
                else None
            )
            rec["primary_care_staff_per_1000"] = (
                rec["primary_care_staff_count"] / (pop * 10)
                if isinstance(rec["primary_care_staff_count"], (int, float))
                else None
            )
            rec["primary_care_beds_per_1000"] = (
                rec["primary_care_beds_count"] / (pop * 10)
                if isinstance(rec["primary_care_beds_count"], (int, float))
                else None
            )
        else:
            rec["primary_care_institutions_per_10000"] = None
            rec["primary_care_staff_per_1000"] = None
            rec["primary_care_beds_per_1000"] = None
        output_rows.append(rec)

    output_path = OUT_DIR / "14_yearbook_primary_care_context_2001_2020.csv"
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(output_rows[0]))
        writer.writeheader()
        writer.writerows(output_rows)

    audit_path = OUT_DIR / "15_yearbook_indicator_audit.csv"
    with audit_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(audit_rows[0]))
        writer.writeheader()
        writer.writerows(audit_rows)

    print(f"Wrote {len(output_rows)} area-year rows: {output_path}")
    print(f"Wrote {len(audit_rows)} indicator definitions: {audit_path}")
    print("First data row:", output_rows[0])
    print("Last data row:", output_rows[-1])


if __name__ == "__main__":
    main()
